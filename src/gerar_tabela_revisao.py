"""
Script de Geracao da Tabela de Revisao Bibliografica — Pos-Triagem

Gera uma tabela-template para extracao manual de dados dos artigos
elegiveis (fase2_decisao == 'incluir'), com campos auto-preenchidos a
partir dos metadados do Lens.org e campos vazios para preenchimento
durante a leitura de texto completo.

Saidas:
    data/tabela_revisao_bibliografica.csv   (formato versatil)
    data/tabela_revisao_bibliografica.xlsx  (formatado para edicao manual:
                                             freeze panes, dropdowns,
                                             auto-filter, cores por categoria)

Protocolo de preenchimento: docs/protocolo_revisao_bibliografica.md

Como usar:
    python src/gerar_tabela_revisao.py
"""

import os
from datetime import datetime
import pandas as pd

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


# ============================================================
# CONFIGURACAO — Caminhos
# ============================================================

PASTA_PROJETO = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
ARQUIVO_ENTRADA = os.path.join(PASTA_PROJETO, "data", "artigos_unicos_triagem.csv")
ARQUIVO_SAIDA_CSV = os.path.join(PASTA_PROJETO, "data", "tabela_revisao_bibliografica.csv")
ARQUIVO_SAIDA_XLSX = os.path.join(PASTA_PROJETO, "data", "tabela_revisao_bibliografica.xlsx")


# ============================================================
# TAXONOMIAS / DROPDOWNS
# ============================================================

# Mapeamento de prioridade do tipo_problema (menor = mais prioritario na ordenacao)
PRIORIDADE_TIPO_PROBLEMA = {
    "backorder_prediction": 1,
    "demand_forecasting": 2,
    "inventory_control": 3,
    "routing_transportation": 4,
    "scheduling_production": 5,
    "supplier_procurement": 6,
    "risk_resilience": 7,
    "sustainability_sc": 8,
    "predictive_maintenance": 9,
    "supply_chain_general": 10,
    "industry40_smart_mfg": 11,
    "qml_method_review": 12,
    "outros": 13,
}

# Dropdowns para validacao Excel
DROPDOWNS = {
    "revisado": ["sim", "nao"],
    "relevancia_final": ["alta", "media", "baixa", "excluir"],
    "problema_validado": list(PRIORIDADE_TIPO_PROBLEMA.keys()),
    "metodo_qml": ["QNN", "VQC", "QSVM", "QRL", "Quantum Kernel", "QRC", "Hybrid", "Other"],
    "dataset_fonte": ["publico", "proprietario", "sintetico", "benchmark"],
    "resultado_qml_vs_classico": ["melhor", "equivalente", "pior", "nao_comparado"],
    "hardware": ["simulator", "real_quantum", "both"],
    "tipo_contribuicao": ["novel_method", "benchmark", "application", "review", "case_study"],
}

# Cores para a coluna tipo_problema_auto (fundo por categoria)
CORES_TIPO_PROBLEMA = {
    # Verde (nucleo CI-2)
    "backorder_prediction":     "C6EFCE",
    "demand_forecasting":       "C6EFCE",
    "inventory_control":        "C6EFCE",
    # Azul claro (SC especifico)
    "routing_transportation":   "BDD7EE",
    "scheduling_production":    "BDD7EE",
    "supplier_procurement":     "BDD7EE",
    "risk_resilience":          "BDD7EE",
    "sustainability_sc":        "BDD7EE",
    "predictive_maintenance":   "BDD7EE",
    # Amarelo (catch-all SC)
    "supply_chain_general":     "FFEB9C",
    # Cinza (baixa prioridade / generico)
    "industry40_smart_mfg":     "D9D9D9",
    "qml_method_review":        "D9D9D9",
    "outros":                   "D9D9D9",
}


# ============================================================
# COLUNAS DA TABELA
# ============================================================

# Ordem: metadados (auto) -> decisao (manual) -> metodo (manual) -> dataset (manual) ->
#        resultados (manual) -> hardware (manual) -> sintese (manual)

COLUNAS_AUTO = [
    "id", "titulo", "autores_completo",
    "ano", "venue", "publication_type",
    "doi", "url",
    "is_open_access", "citing_works", "citacoes_por_ano",
    "tipo_problema_auto",
]

COLUNAS_MANUAIS = [
    # Decisao do revisor
    "revisado", "relevancia_final", "problema_validado",
    # Metodo
    "metodo_qml", "metodo_qml_detalhe",
    # Baseline
    "baseline_classico",
    # Dataset
    "dataset_nome", "dataset_tamanho", "dataset_fonte",
    # Resultados
    "metricas", "resultado_qml_vs_classico", "diferenca_percentual",
    # Hardware
    "hardware", "hardware_detalhe", "n_qubits",
    # Sintese
    "limitacoes", "tipo_contribuicao", "contribuicao_para_sc",
    "notas",
]

COLUNAS_TODAS = COLUNAS_AUTO + COLUNAS_MANUAIS


# ============================================================
# 1. CARREGAR E FILTRAR
# ============================================================

def carregar_e_filtrar():
    if not os.path.exists(ARQUIVO_ENTRADA):
        raise FileNotFoundError(
            f"Arquivo nao encontrado: {ARQUIVO_ENTRADA}\n"
            "Execute primeiro: python src/triagem_artigos.py && python src/classificar_problema.py"
        )
    df = pd.read_csv(ARQUIVO_ENTRADA)
    print(f"Lidos {len(df)} artigos do CSV de triagem")

    if "fase2_decisao" not in df.columns:
        raise ValueError("Coluna 'fase2_decisao' nao existe. Rode triagem_artigos.py primeiro.")
    if "tipo_problema" not in df.columns:
        raise ValueError("Coluna 'tipo_problema' nao existe. Rode classificar_problema.py primeiro.")

    elegiveis = df[df["fase2_decisao"] == "incluir"].copy()
    print(f"Filtrados {len(elegiveis)} artigos elegiveis (fase2_decisao == 'incluir')")
    return elegiveis


# ============================================================
# 2. MONTAR TABELA
# ============================================================

def montar_tabela(df_elegiveis):
    out = pd.DataFrame(index=df_elegiveis.index)

    # Auto-preenchidas
    out["id"] = df_elegiveis["Lens ID"]
    out["titulo"] = df_elegiveis["Title"]
    out["autores_completo"] = df_elegiveis["Author/s"]
    out["ano"] = df_elegiveis["Publication Year"]
    out["venue"] = df_elegiveis["Source Title"]
    out["publication_type"] = df_elegiveis["Publication Type"]
    out["doi"] = df_elegiveis["DOI"]
    out["url"] = df_elegiveis["External URL"]
    out["is_open_access"] = df_elegiveis["Is Open Access"]
    out["citing_works"] = df_elegiveis["Citing Works Count"]

    # Metrica ajustada pela idade do artigo: citacoes / anos desde a publicacao.
    # Corrige vies temporal — artigos recentes com poucas citacoes nao sao
    # penalizados apenas por serem jovens, e artigos antigos nao dominam por tempo.
    ano_atual = datetime.now().year
    anos_desde_pub = (ano_atual - pd.to_numeric(df_elegiveis["Publication Year"], errors="coerce")).clip(lower=1)
    citacoes = pd.to_numeric(df_elegiveis["Citing Works Count"], errors="coerce").fillna(0)
    out["citacoes_por_ano"] = (citacoes / anos_desde_pub).round(2)

    out["tipo_problema_auto"] = df_elegiveis["tipo_problema"]

    # Manuais (vazias)
    for c in COLUNAS_MANUAIS:
        out[c] = ""

    # Ordenacao
    out["_prioridade"] = out["tipo_problema_auto"].map(PRIORIDADE_TIPO_PROBLEMA).fillna(99)
    out = out.sort_values(
        by=["_prioridade", "ano", "citing_works"],
        ascending=[True, False, False],
        na_position="last",
    )
    out = out.drop(columns=["_prioridade"]).reset_index(drop=True)

    # Garantir ordem das colunas
    out = out[COLUNAS_TODAS]

    return out


# ============================================================
# 3. EXPORTAR CSV
# ============================================================

def exportar_csv(df):
    df.to_csv(ARQUIVO_SAIDA_CSV, index=False)
    print(f"CSV exportado: {ARQUIVO_SAIDA_CSV}")


# ============================================================
# 4. EXPORTAR XLSX FORMATADO
# ============================================================

def exportar_xlsx(df):
    wb = Workbook()
    ws = wb.active
    ws.title = "revisao_bibliografica"

    # Header
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="404040")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, coluna in enumerate(COLUNAS_TODAS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=coluna)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Dados
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        for col_idx, valor in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=valor if pd.notna(valor) else "")

    # Cor de fundo por categoria em tipo_problema_auto
    col_tipo_problema = COLUNAS_TODAS.index("tipo_problema_auto") + 1
    for row_idx in range(2, len(df) + 2):
        tipo = ws.cell(row=row_idx, column=col_tipo_problema).value
        cor = CORES_TIPO_PROBLEMA.get(tipo)
        if cor:
            ws.cell(row=row_idx, column=col_tipo_problema).fill = PatternFill("solid", fgColor=cor)

    # Freeze panes: linha 1 (header) + colunas A e B (id, titulo)
    ws.freeze_panes = "C2"

    # Auto-filter em todo o range
    ultima_col = get_column_letter(len(COLUNAS_TODAS))
    ws.auto_filter.ref = f"A1:{ultima_col}{len(df) + 1}"

    # Largura das colunas
    larguras = {
        "id": 22, "titulo": 50, "autores_completo": 40,
        "ano": 8, "venue": 30, "publication_type": 18,
        "doi": 28, "url": 30,
        "is_open_access": 12, "citing_works": 10, "citacoes_por_ano": 12,
        "tipo_problema_auto": 22,
        "revisado": 10, "relevancia_final": 14, "problema_validado": 22,
        "metodo_qml": 14, "metodo_qml_detalhe": 30,
        "baseline_classico": 22,
        "dataset_nome": 22, "dataset_tamanho": 12, "dataset_fonte": 14,
        "metricas": 30, "resultado_qml_vs_classico": 18, "diferenca_percentual": 16,
        "hardware": 12, "hardware_detalhe": 22, "n_qubits": 10,
        "limitacoes": 30, "tipo_contribuicao": 16, "contribuicao_para_sc": 30,
        "notas": 30,
    }
    for col_idx, coluna in enumerate(COLUNAS_TODAS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = larguras.get(coluna, 16)

    ws.row_dimensions[1].height = 32

    # Data validations (dropdowns)
    # Observacao: data validation com formula1='"a,b,c"' (limite de 255 chars por formula).
    # problema_validado tem 13 valores — cabe na string.
    for coluna, opcoes in DROPDOWNS.items():
        col_idx = COLUNAS_TODAS.index(coluna) + 1
        letra = get_column_letter(col_idx)
        formula = '"' + ",".join(opcoes) + '"'
        dv = DataValidation(type="list", formula1=formula, allow_blank=True)
        dv.error = "Valor invalido — use um dos itens do dropdown"
        dv.errorTitle = "Valor invalido"
        dv.prompt = "Selecione um valor da lista"
        dv.promptTitle = coluna
        dv.add(f"{letra}2:{letra}{len(df) + 1}")
        ws.add_data_validation(dv)

    wb.save(ARQUIVO_SAIDA_XLSX)
    print(f"XLSX exportado: {ARQUIVO_SAIDA_XLSX}")


# ============================================================
# 5. RESUMO
# ============================================================

def imprimir_resumo(df):
    print("\n" + "=" * 60)
    print("  RESUMO — Tabela de Revisao Bibliografica")
    print("=" * 60)
    print(f"  Total de artigos: {len(df)}")
    print(f"  Colunas auto-preenchidas: {len(COLUNAS_AUTO)}")
    print(f"  Colunas para extracao manual: {len(COLUNAS_MANUAIS)}")
    print(f"  Total de colunas: {len(COLUNAS_TODAS)}\n")

    print("  Distribuicao por tipo_problema_auto:")
    dist = df["tipo_problema_auto"].value_counts()
    for cat in PRIORIDADE_TIPO_PROBLEMA.keys():
        n = int(dist.get(cat, 0))
        if n > 0:
            print(f"    {cat:28s}: {n:3d}")
    print()


def main():
    elegiveis = carregar_e_filtrar()
    tabela = montar_tabela(elegiveis)
    exportar_csv(tabela)
    exportar_xlsx(tabela)
    imprimir_resumo(tabela)
    print("=" * 60)
    print("  Tabela gerada com sucesso!")
    print("=" * 60)
    print("\n  Proximos passos:")
    print(f"    1. Abrir {os.path.basename(ARQUIVO_SAIDA_XLSX)} no Excel/LibreOffice")
    print("    2. Preencher colunas manuais durante leitura de texto completo")
    print("    3. Consultar docs/protocolo_revisao_bibliografica.md para convencoes\n")


if __name__ == "__main__":
    main()
